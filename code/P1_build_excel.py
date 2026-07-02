"""
build_excel.py — SaaS/Telecom Churn Analysis Excel Deliverable
==============================================================
Builds a 7-sheet professional workbook structured as a consultant would
present findings to a client: cover → situation → findings → recommendation.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path("outputs")

# ── Colour tokens ─────────────────────────────────────────────────────────────
NAVY     = "1A3C5E"
CORAL    = "C0392B"
AMBER    = "E67E22"
GREEN    = "27AE60"
SKY      = "2980B9"
OFFWHITE = "FAFAFA"
LTGREY   = "ECF0F1"
MIDGREY  = "95A5A6"
DGREY    = "2C3E50"
WHITE    = "FFFFFF"
AMBER_BG = "FEF9E7"
RED_BG   = "FDEDEC"
GREEN_BG = "EAFAF1"

def fill(hex_):    return PatternFill("solid", fgColor=hex_)
def font(size=11, color="000000", bold=False, italic=False):
    return Font(name="Arial", size=size, color=color, bold=bold, italic=italic)

thin   = Side(style="thin",   color="BDC3C7")
medium = Side(style="medium", color=NAVY)

def border_all(cell):
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def border_bottom(cell):
    cell.border = Border(bottom=Side(style="medium", color=NAVY))

def write(ws, row, col, value, sz=11, color="000000", bold=False,
          italic=False, bg=None, align="center", wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(sz, color, bold, italic)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    border_all(c)
    if num_fmt:
        c.number_format = num_fmt
    return c

def section_header(ws, row, text, cols=8, bg=NAVY, color=WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = font(11, color, bold=True)
    c.fill  = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def title_banner(ws, title, subtitle, cols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    t = ws.cell(row=1, column=1, value=title)
    t.font  = font(16, WHITE, bold=True)
    t.fill  = fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font  = font(10, DGREY, italic=True)
    s.fill  = fill(LTGREY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 18

def col_header(ws, row, headers, widths, bg=NAVY, fg=WHITE):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font  = font(10, fg, bold=True)
        c.fill  = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_all(c)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

def spacer(ws, row, cols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 8

# ── Load outputs ──────────────────────────────────────────────────────────────
cohort_df  = pd.read_csv(OUT / "cohort_summary.csv")
contract_df = pd.read_csv(OUT / "contract_summary.csv")
service_df  = pd.read_csv(OUT / "service_adoption_summary.csv")
risk_df     = pd.read_csv(OUT / "risk_segment_summary.csv")
fi_df       = pd.read_csv(OUT / "model_feature_importance.csv")

wb = Workbook()


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER / SITUATION
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Cover & Situation"
ws1.sheet_view.showGridLines = False

title_banner(ws1,
    "Customer Churn Analysis — Telecom Subscription Business",
    "Dataset: IBM Telco Customer Churn (7,043 customers, 21 variables) · Analyst: Vaishnavi Bhor")

# KPI strip — row 4/5/6
kpis = [
    ("Total Customers", "7,043",    "Q3 snapshot, California",        SKY),
    ("Overall Churn",   "26.5%",    "1,869 churned customers",        CORAL),
    ("Early Cohort\nChurn", "52.9%","0–6 months tenure",              CORAL),
    ("Established\nChurn",  "14.0%","25–72 months tenure",            GREEN),
    ("Model ROC-AUC",   "0.842",    "Gradient Boosting, 5-fold CV",   SKY),
    ("Highest-Risk\nSegment", "65.8%","989 customers at risk",         CORAL),
]
for i, (label, val, note, color) in enumerate(kpis, 1):
    ws1.column_dimensions[get_column_letter(i)].width = 20
    c1 = ws1.cell(row=4, column=i, value=label)
    c1.font  = font(9, WHITE, bold=True)
    c1.fill  = fill(color)
    c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[4].height = 30

    c2 = ws1.cell(row=5, column=i, value=val)
    c2.font  = font(18, color, bold=True)
    c2.fill  = fill(OFFWHITE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[5].height = 38

    c3 = ws1.cell(row=6, column=i, value=note)
    c3.font  = font(8, MIDGREY, italic=True)
    c3.fill  = fill(OFFWHITE)
    c3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[6].height = 24

spacer(ws1, 7)

# Situation / client brief
section_header(ws1, 8, "SITUATION — WHAT WAS ASKED")
rows_sit = [
    ("Business context", "Telecom subscription company serving ~7,000 customers in California with home phone and internet services. Churn is the primary profitability risk."),
    ("Analytical question", "WHERE is churn concentrated? WHAT structural factors drive it? WHICH customer segments are highest risk, and WHY?"),
    ("What this is NOT", "This is not a post-implementation result report. These are analytical findings and recommendations derived from real customer data. Implementation and outcome measurement would be the client's next step."),
    ("Data source", "IBM Telco Customer Churn dataset (public, IBM Cognos Analytics sample). Real customer records. 7,043 rows × 21 columns. No synthetic data."),
]
for i, (label, desc) in enumerate(rows_sit, 9):
    ws1.row_dimensions[i].height = 32
    c1 = ws1.cell(row=i, column=1, value=label)
    c1.font  = font(10, NAVY, bold=True)
    c1.fill  = fill(LTGREY if i % 2 == 0 else OFFWHITE)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    border_all(c1)
    ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    c2 = ws1.cell(row=i, column=2, value=desc)
    c2.font  = font(10)
    c2.fill  = fill(LTGREY if i % 2 == 0 else OFFWHITE)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    border_all(c2)

spacer(ws1, 13)

# Issue tree overview
section_header(ws1, 14, "ANALYTICAL FRAMEWORK — McKINSEY ISSUE TREE")
branches = [
    ("Branch 1", "Tenure Cohort Analysis",     "Is churn a new-customer problem or evenly distributed across lifecycle?",  "Sheet: Cohort Analysis"),
    ("Branch 2", "Contract Type",              "Does contract length (commitment level) predict churn?",                    "Sheet: Contract Analysis"),
    ("Branch 3", "Service Adoption Depth",     "Do customers with more services churn less? (stickiness hypothesis)",        "Sheet: Service Depth"),
    ("Branch 4", "Service Differentiators",    "Which specific services are under-adopted by churners vs retained?",         "Sheet: Service Analysis"),
    ("Branch 5", "Risk Segment Isolation",     "Can we isolate a highest-risk segment from combined factor filters?",        "Sheet: Risk Segment"),
    ("Branch 6", "Predictive Model (GB)",      "Which variables have greatest predictive power? (ROC-AUC = 0.842)",         "Sheet: Model Output"),
]
col_header(ws1, 15, ["Branch", "Name", "Analytical Question", "Output"], [10, 24, 48, 22])
for i, (branch, name, q, output) in enumerate(branches, 16):
    ws1.row_dimensions[i].height = 24
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    write(ws1, i, 1, branch,  10, CORAL, bold=True, bg=bg)
    write(ws1, i, 2, name,    10, DGREY, bold=True, bg=bg, align="left")
    write(ws1, i, 3, q,       10, bg=bg, align="left")
    write(ws1, i, 4, output,  9, SKY, bg=bg, align="left", italic=True)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — COHORT ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Cohort Analysis")
ws2.sheet_view.showGridLines = False
title_banner(ws2, "Branch 1: Churn by Customer Tenure Cohort",
             "Question: Is churn evenly distributed across the customer lifecycle, or concentrated in a specific window?")

section_header(ws2, 4, "COHORT SUMMARY TABLE")
col_header(ws2, 5,
    ["Tenure Cohort", "Customers", "Churned", "Retained", "Churn Rate", "Avg Monthly (£)", "Avg Value Services Adopted"],
    [22, 14, 12, 12, 14, 18, 28])

COHORT_ORDER = ["0–6 months\n(Early)", "7–24 months\n(Developing)", "25–72 months\n(Established)"]
cohort_df["tenure_cohort"] = pd.Categorical(cohort_df["tenure_cohort"], COHORT_ORDER, ordered=True)
cohort_df = cohort_df.sort_values("tenure_cohort").reset_index(drop=True)

highlight_colors = [RED_BG, AMBER_BG, GREEN_BG]
for i, (_, row) in enumerate(cohort_df.iterrows()):
    bg = highlight_colors[i]
    r = i + 6
    ws2.row_dimensions[r].height = 26
    write(ws2, r, 1, row["tenure_cohort"].replace("\n", " "), 10, DGREY, bold=True, bg=bg, align="left")
    write(ws2, r, 2, row["total"],                10, bg=bg, num_fmt="#,##0")
    write(ws2, r, 3, row["churned"],              10, bg=bg, num_fmt="#,##0")
    write(ws2, r, 4, row["total"] - row["churned"], 10, bg=bg, num_fmt="#,##0")
    churn_col = CORAL if row["churn_rate"] > 0.40 else AMBER if row["churn_rate"] > 0.20 else GREEN
    write(ws2, r, 5, row["churn_rate"],           11, churn_col, bold=True, bg=bg, num_fmt="0.0%")
    write(ws2, r, 6, row["avg_monthly"],          10, bg=bg, num_fmt="£#,##0.00")
    write(ws2, r, 7, row["avg_value_services"],   10, bg=bg, num_fmt="0.00")

spacer(ws2, 9)

section_header(ws2, 10, "FINDING & INTERPRETATION")
findings = [
    ("Finding",      "Churn is front-loaded. Customers in their first 6 months churn at 52.9% — nearly 4x the rate of established customers (14.0%)."),
    ("Implication",  "The customer onboarding window is the highest-leverage intervention point. Retention economics favour early investment: a customer saved in month 3 has a much longer potential LTV than one saved in month 30."),
    ("Mechanism",    "Early customers have lower service adoption (avg 1.96 value services vs 4.19 for established). Low adoption = low switching cost = low perceived value. The customer hasn't yet embedded the product into their workflow."),
    ("What to test", "Interventions targeting the 0–6 month cohort: contract upgrade offers, proactive value service bundling, and assigned account touches at Day 7, Day 30, and Day 90."),
]
for i, (label, desc) in enumerate(findings, 11):
    ws2.row_dimensions[i].height = 32
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    c1 = ws2.cell(row=i, column=1, value=label)
    c1.font  = font(10, NAVY, bold=True)
    c1.fill  = fill(bg)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    border_all(c1)
    ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
    c2 = ws2.cell(row=i, column=2, value=desc)
    c2.font  = font(10)
    c2.fill  = fill(bg)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    border_all(c2)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — CONTRACT ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Contract Analysis")
ws3.sheet_view.showGridLines = False
title_banner(ws3, "Branch 2: Contract Type as Commitment Signal",
             "Question: Does contract length predict churn? (Commitment = switching cost proxy)")

section_header(ws3, 4, "CHURN RATE BY COHORT × CONTRACT TYPE")
col_header(ws3, 5,
    ["Tenure Cohort", "Contract Type", "Customers", "Churned", "Churn Rate", "Risk Level"],
    [24, 20, 14, 12, 14, 14])

contract_df["tenure_cohort"] = contract_df["tenure_cohort"].str.replace("\n", " ")
contract_df = contract_df.sort_values(["tenure_cohort", "churn_rate"], ascending=[True, False]).reset_index(drop=True)

for i, (_, row) in enumerate(contract_df.iterrows()):
    r = i + 6
    ws3.row_dimensions[r].height = 24
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    risk = "HIGH" if row["churn_rate"] > 0.40 else "MEDIUM" if row["churn_rate"] > 0.15 else "LOW"
    risk_col = CORAL if risk == "HIGH" else AMBER if risk == "MEDIUM" else GREEN
    write(ws3, r, 1, row["tenure_cohort"],  10, DGREY, bg=bg, align="left")
    write(ws3, r, 2, row["Contract"],       10, bg=bg, align="left")
    churned_n = int(round(row["customers"] * row["churn_rate"]))
    write(ws3, r, 3, row["customers"],  10, bg=bg, num_fmt="#,##0")
    write(ws3, r, 4, churned_n,         10, bg=bg, num_fmt="#,##0")
    write(ws3, r, 5, row["churn_rate"], 11, risk_col, bold=True, bg=bg, num_fmt="0.0%")
    write(ws3, r, 6, risk, 10, risk_col, bold=True, bg=bg)

spacer(ws3, 15)

section_header(ws3, 16, "FINDING & INTERPRETATION")
findings3 = [
    ("Core finding", "Contract type is the single strongest structural churn lever, confirmed by both the descriptive analysis and the predictive model (35% feature importance — #1 of all 19 variables)."),
    ("M2M vs Annual", "Month-to-month customers churn at 42.7%. One-year contract customers: 11.3%. Two-year: 2.8%. This is not a marginal difference — it is a structural one."),
    ("Why it matters", "Contract length is a proxy for commitment and switching cost. A customer on a two-year contract has already made a deliberate decision to stay. That decision should be accelerated at onboarding."),
    ("Recommendation", "Offer a contract upgrade incentive (e.g., free month, discounted add-on) at the 30-day and 90-day marks for month-to-month customers — particularly those in the early cohort who are highest risk."),
]
for i, (label, desc) in enumerate(findings3, 17):
    ws3.row_dimensions[i].height = 32
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    c1 = ws3.cell(row=i, column=1, value=label)
    c1.font = font(10, NAVY, bold=True)
    c1.fill = fill(bg)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    border_all(c1)
    ws3.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    c2 = ws3.cell(row=i, column=2, value=desc)
    c2.font = font(10)
    c2.fill = fill(bg)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    border_all(c2)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SERVICE ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Service Analysis")
ws4.sheet_view.showGridLines = False
title_banner(ws4, "Branch 4: Service Differentiators — Which Services Drive Retention?",
             "Chi-square tested: services where churners and retained customers differ significantly in adoption rate")

section_header(ws4, 4, "SERVICE ADOPTION RATE: CHURNED vs RETAINED CUSTOMERS")
col_header(ws4, 5,
    ["Service", "Churned\nAdoption %", "Retained\nAdoption %", "Gap (pp)", "Stat. Significant?", "Role in Retention"],
    [22, 18, 18, 14, 20, 40])

service_df_sorted = service_df.sort_values("gap", ascending=False)
for i, (_, row) in enumerate(service_df_sorted.iterrows()):
    r = i + 6
    ws4.row_dimensions[r].height = 26
    bg = GREEN_BG if row["gap"] > 0.10 else RED_BG if row["gap"] < -0.05 else OFFWHITE
    gap_col = GREEN if row["gap"] > 0.10 else CORAL if row["gap"] < -0.05 else MIDGREY
    sig_text = "YES ★" if row["significant"] else "No"
    sig_col  = GREEN if row["significant"] else MIDGREY

    role = {
        "Online Security":    "★ HIGH — security services create perceived indispensability",
        "Tech Support":       "★ HIGH — support relationship increases perceived value",
        "Online Backup":      "MEDIUM — data dependency creates mild switching cost",
        "Device Protection":  "MEDIUM — asset protection creates commitment",
        "Streaming TV":       "LOW — entertainment add-on, not workflow dependency",
        "Streaming Movies":   "LOW — entertainment add-on, not workflow dependency",
        "Multiple Lines":     "NEUTRAL — household size signal, not a retention lever",
        "Phone Service":      "NEUTRAL — base service, near-universal adoption",
    }.get(row["service"], "")

    write(ws4, r, 1, row["service"],             10, DGREY, bold=True, bg=bg, align="left")
    write(ws4, r, 2, row["churned_adoption"],    10, bg=bg, num_fmt="0.0%")
    write(ws4, r, 3, row["retained_adoption"],   10, bg=bg, num_fmt="0.0%")
    write(ws4, r, 4, row["gap"],                 10, gap_col, bold=True, bg=bg, num_fmt="+0.0%;-0.0%;0.0%")
    write(ws4, r, 5, sig_text,                   10, sig_col, bold=True, bg=bg)
    write(ws4, r, 6, role,                        9, DGREY, bg=bg, align="left", italic=True)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — RISK SEGMENT
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Risk Segment")
ws5.sheet_view.showGridLines = False
title_banner(ws5, "Branch 5: Highest-Risk Customer Segment — Build-up & Isolation",
             "Progressive filter logic: combining factors to isolate the highest-risk customer cluster")

section_header(ws5, 4, "RISK SEGMENT WATERFALL")
col_header(ws5, 5,
    ["Segment", "Filter Applied", "Customers", "Churned", "Churn Rate", "Avg Monthly (£)", "Risk Level"],
    [30, 34, 14, 12, 14, 18, 14])

for i, (_, row) in enumerate(risk_df.iterrows()):
    r = i + 6
    ws5.row_dimensions[r].height = 26
    bg = RED_BG if row["churn_rate"] > 0.55 else AMBER_BG if row["churn_rate"] > 0.35 else OFFWHITE
    risk = "CRITICAL" if row["churn_rate"] > 0.55 else "HIGH" if row["churn_rate"] > 0.35 else "BASELINE"
    risk_col = CORAL if risk == "CRITICAL" else AMBER if risk == "HIGH" else MIDGREY

    write(ws5, r, 1, row["segment"],             10, DGREY, bold=(i==len(risk_df)-1), bg=bg, align="left")
    write(ws5, r, 2, row["filter"],              9, DGREY, bg=bg, align="left", italic=True)
    write(ws5, r, 3, row["customers"],           10, bg=bg, num_fmt="#,##0")
    write(ws5, r, 4, row["churned"],             10, bg=bg, num_fmt="#,##0")
    write(ws5, r, 5, row["churn_rate"],          11, risk_col, bold=True, bg=bg, num_fmt="0.0%")
    write(ws5, r, 6, row["avg_monthly"],         10, bg=bg, num_fmt="£#,##0.00")
    write(ws5, r, 7, risk,                       10, risk_col, bold=True, bg=bg)

spacer(ws5, 12)
section_header(ws5, 13, "SEGMENT PROFILE & RECOMMENDED ACTION")
profile_rows = [
    ("Who they are",    "Month-to-month contract, Fiber optic internet, Electronic check payment, no Online Security, no Tech Support. 989 customers. Avg £85.23/month."),
    ("Why they churn",  "Every factor compounds: no commitment signal (M2M), highest-cost service (Fiber), frictionless payment (Elec. check makes cancellation easy), zero value-add services (no perceived indispensability)."),
    ("Revenue at risk", "989 customers × £85.23/month = approximately £84,242/month in MRR at high churn risk. Even recovering 20% of this segment represents meaningful revenue protection."),
    ("Recommended action","Proactive outreach programme: (1) contract upgrade offer with free OnlineSecurity trial for 3 months, (2) autopay switch incentive (e.g. 5% discount), (3) assigned customer success touch at months 1 and 3. These are hypotheses to test — A/B test design required before rollout."),
    ("What this is not", "This is a model-derived prioritisation, not a guarantee. Correlation is not causation. The recommended actions are structured hypotheses that should be validated through controlled testing before full deployment."),
]
for i, (label, desc) in enumerate(profile_rows, 14):
    ws5.row_dimensions[i].height = 34
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    c1 = ws5.cell(row=i, column=1, value=label)
    c1.font = font(10, NAVY, bold=True)
    c1.fill = fill(bg)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    border_all(c1)
    ws5.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
    c2 = ws5.cell(row=i, column=2, value=desc)
    c2.font = font(10)
    c2.fill = fill(bg)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    border_all(c2)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 — MODEL OUTPUT
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Model Output")
ws6.sheet_view.showGridLines = False
title_banner(ws6, "Branch 6: Gradient Boosting Model — Churn Prediction & Feature Importance",
             "Model: Gradient Boosting Classifier · ROC-AUC: 0.842 · 5-fold CV: 0.842 ± 0.008")

# Model performance block
section_header(ws6, 4, "MODEL PERFORMANCE SUMMARY")
col_header(ws6, 5,
    ["Metric", "Value", "Interpretation"],
    [24, 16, 55])
perf_rows = [
    ("Algorithm",         "Gradient Boosting Classifier",   "Ensemble method; handles non-linear relationships and feature interactions well."),
    ("Training split",    "75% train / 25% test (stratified)","Stratified split preserves class imbalance (73.5% / 26.5%) in both sets."),
    ("ROC-AUC (test set)","0.842",                           "Probability that the model ranks a random churner above a random non-churner. 0.84 is strong for a business churn model."),
    ("5-fold CV AUC",     "0.842 ± 0.008",                  "Consistent across folds — the model is not overfitting. Generalises well to unseen customers."),
    ("Precision (churn)", "0.67",                            "67% of customers the model flags as churners will actually churn. Acceptable for prioritisation use."),
    ("Recall (churn)",    "0.48",                            "The model captures 48% of actual churners. Trade-off: conservative flagging reduces false alarms."),
]
for i, (m, v, interp) in enumerate(perf_rows, 6):
    ws6.row_dimensions[i].height = 26
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    write(ws6, i, 1, m,     10, NAVY, bold=True, bg=bg, align="left")
    write(ws6, i, 2, v,     10, DGREY, bold=True, bg=bg)
    ws6.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    c = ws6.cell(row=i, column=3, value=interp)
    c.font = font(10)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    border_all(c)

spacer(ws6, 12)

section_header(ws6, 13, "FEATURE IMPORTANCE (Top 10 Churn Predictors)")
col_header(ws6, 14,
    ["Rank", "Variable", "Readable Label", "Importance %", "Business Interpretation"],
    [8, 22, 26, 16, 45])

fi_df["label"] = fi_df.get("label", fi_df["feature"])
for i, (_, row) in enumerate(fi_df.head(10).iterrows()):
    r = i + 15
    ws6.row_dimensions[r].height = 26
    bg = RED_BG if row["importance"] > 0.15 else AMBER_BG if row["importance"] > 0.08 else OFFWHITE
    imp_col = CORAL if row["importance"] > 0.15 else AMBER if row["importance"] > 0.08 else SKY
    interp = {
        "Contract":        "Contract type is the #1 predictor. M2M customers are structurally more likely to leave.",
        "TotalCharges":    "Proxy for tenure × spend. Low total charges = early-stage customer at higher risk.",
        "MonthlyCharges":  "Higher charges correlate with more services (stickier) but also more price sensitivity.",
        "tenure":          "Time as a customer. Longer tenure = lower churn risk. Confirms cohort finding.",
        "OnlineSecurity":  "Key protective service. Churners adopt it at 15.8% vs 33.3% for retained.",
        "TechSupport":     "Support relationship = perceived value. Churners adopt at 16.6% vs 33.5%.",
        "InternetService": "Fiber optic customers churn at 42% vs DSL at 19%. Service quality signal.",
        "PaperlessBilling":"Electronic billing correlates with electronic check — both signal lower friction to cancel.",
        "PaymentMethod":   "Electronic check = highest churn (45%). Autopay methods = lowest churn (15–17%).",
        "MultipleLines":   "Lower importance — household size signal more than a stickiness lever.",
    }.get(row["feature"], "")

    write(ws6, r, 1, i + 1,              10, DGREY, bg=bg)
    write(ws6, r, 2, row["feature"],     10, DGREY, bg=bg, align="left")
    write(ws6, r, 3, row.get("label", row["feature"]), 10, DGREY, bold=True, bg=bg, align="left")
    write(ws6, r, 4, row["importance"],  11, imp_col, bold=True, bg=bg, num_fmt="0.0%")
    ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=5)
    c = ws6.cell(row=r, column=5, value=interp)
    c.font = font(9, DGREY)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    border_all(c)
    ws6.column_dimensions["E"].width = 55


# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 — RECOMMENDATIONS & METHODOLOGY
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Recommendations")
ws7.sheet_view.showGridLines = False
title_banner(ws7, "Recommendations & Honest Validation Notes",
             "Structured recommendations + transparency on methodology, assumptions, and limitations")

section_header(ws7, 4, "PRIORITY RECOMMENDATIONS (Hypotheses to Test — Not Implemented Results)")
col_header(ws7, 5,
    ["Priority", "Recommendation", "Target Segment", "Mechanism", "How to Test"],
    [10, 35, 22, 30, 35])

recs = [
    ("1 — HIGH",   "Introduce contract upgrade offer at Day 30 and Day 90 for month-to-month customers",
                    "M2M, 0–6 month cohort",
                    "Increases switching cost and commitment signal. Contract type = #1 churn predictor (35% model importance).",
                    "A/B test: 50% of M2M new customers receive offer at Day 30. Measure 90-day churn rate vs control."),
    ("2 — HIGH",   "Bundle OnlineSecurity or TechSupport free trial (3 months) at onboarding",
                    "All new customers, especially M2M",
                    "Service adoption gap: churners adopt security at 15.8% vs retained at 33.3%. Trial removes friction to adoption.",
                    "A/B test: 50% of new signups receive free trial. Measure Day-90 activation rate and 6-month churn rate."),
    ("3 — MEDIUM", "Proactive autopay migration campaign for electronic check customers",
                    "Electronic check segment (2,365 customers, 45.3% churn)",
                    "Electronic check = highest churn payment method. Autopay reduces cancellation friction AND improves payment reliability.",
                    "Outbound campaign offering 5% discount for autopay switch. Track conversion rate and 90-day churn change."),
    ("4 — MEDIUM", "Assign dedicated customer success touches to highest-risk segment",
                    "989-customer CRITICAL segment (M2M + Fiber + Elec. check + zero value services)",
                    "Human intervention before churn decision is made. Proactive outreach at Month 1 and Month 3.",
                    "Pilot with 50% of segment. Measure contact rate, upgrade conversion, and 6-month retention vs control."),
]
for i, (pri, rec, seg, mech, test) in enumerate(recs, 6):
    ws7.row_dimensions[i].height = 48
    bg = RED_BG if "HIGH" in pri else AMBER_BG
    pri_col = CORAL if "HIGH" in pri else AMBER
    write(ws7, i, 1, pri,   10, pri_col, bold=True, bg=bg)
    write(ws7, i, 2, rec,   10, DGREY, bg=bg, align="left", wrap=True)
    write(ws7, i, 3, seg,   9,  DGREY, bg=bg, align="left", wrap=True)
    write(ws7, i, 4, mech,  9,  DGREY, bg=bg, align="left", wrap=True)
    write(ws7, i, 5, test,  9,  SKY,   bg=bg, align="left", wrap=True, italic=True)

spacer(ws7, 10)
section_header(ws7, 11, "METHODOLOGY & HONEST VALIDATION NOTES")
method_rows = [
    ("Dataset",          "IBM Telco Customer Churn — public dataset, IBM Cognos Analytics sample. Real customer records, not synthetic. 7,043 customers × 21 variables. Publicly accessible via IBM GitHub repository."),
    ("Analytical method","Issue-tree structured analysis (6 branches). Descriptive statistics, chi-square tests for service significance, cohort segmentation, risk segment isolation via progressive filtering, Gradient Boosting predictive model."),
    ("Model validation", "75/25 stratified train-test split + 5-fold cross-validation. ROC-AUC 0.842 with CV standard deviation of ±0.008 — model is stable and not overfitting."),
    ("Correlation ≠ causation","All findings are associative, not causal. A customer on electronic check payment may churn for reasons correlated with electronic check but not caused by it. Controlled A/B testing is required before attributing causal effect to any intervention."),
    ("Recommendations",  "All four recommendations are analytical hypotheses. They are grounded in the data but have not been implemented or measured. Results would depend on execution quality, market conditions, and customer context at time of deployment."),
    ("Limitations",      "Dataset represents a single time-period snapshot (Q3). No longitudinal view of individual customer behaviour. Churn 'reasons' are not recorded. External factors (competitor pricing, market events) are not controlled for."),
    ("What I'd do next", "With implementation access: (1) add a churn reason survey at cancellation, (2) track cohort survival curves over time, (3) build a customer health score using real-time feature usage signals, not just billing and contract data."),
]
for i, (label, desc) in enumerate(method_rows, 12):
    ws7.row_dimensions[i].height = 36
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    c1 = ws7.cell(row=i, column=1, value=label)
    c1.font = font(10, NAVY, bold=True)
    c1.fill = fill(bg)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    border_all(c1)
    ws7.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
    c2 = ws7.cell(row=i, column=2, value=desc)
    c2.font = font(10)
    c2.fill = fill(bg)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    border_all(c2)

spacer(ws7, 19)
section_header(ws7, 20, "ANALYST")
ws7.merge_cells(start_row=21, start_column=1, end_row=21, end_column=5)
c = ws7.cell(row=21, column=1,
             value="Vaishnavi Bhor · MSc Business Analytics, University of Manchester · "
                   "vbhor207@gmail.com · linkedin.com/in/vaishnavi-bhor-business-analyst · vbho.github.io/portfolio")
c.font = font(10, SKY)
c.fill = fill(OFFWHITE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
border_all(c)
ws7.row_dimensions[21].height = 22


# ── Save ──────────────────────────────────────────────────────────────────────
outpath = OUT / "Churn_Analysis_Telecom_VaishnaviBhor.xlsx"
wb.save(outpath)
print(f"\n✓  Workbook saved → {outpath}")
print(f"   Sheets: {len(wb.sheetnames)}")
